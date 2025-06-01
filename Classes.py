import openpyxl


class Coor(object):
    #pass
    X:float
    Y:float
    def __init__(self):
        self.X=0
        self.Y=0

    def __init__(self,x,y):
        self.X=x
        self.Y=y
    def Print(self):
        print(str(self.X)+", " + str(self.Y))
        

class Wall(object):
    name: str # Имя стенки
    name_wall:str
    name_founding:str
    count:int
    foundation_base:float # Отметка подошвы фундамента
    leght: float # Длина секции, м - L
    height_end: float # Высота в конце стенки - H_кон
    height_start: float # Высота в начале стенки - H_нач
    foundation_width: float #Ширина ф-та - В
    top_wall_width: float #ширина стены наерху - в2
    edge_distance: float #расстояние от стены до края - в3
    bottom_wall_width: float #ширина стены внизу - в4
    t1: float #толщина перекрытия 1 у стены 
    t2: float #толщина перекрытия 1 у насыпи
    t3: float #толщина перекрытия 2 у стены
    t4: float #олщина перекрытия 2 у насыпи
    V1: float #объем ростверка подпорной стены, м3
    V2: float #объем стеновой части подпорной стены, м3
    tek_start_x:float
    tek_start_y:float
    tek_end_x:float
    tek_end_y:float
        
    def SetData(self, sheet,read_column):#filename, sheetName
       # workbook = openpyxl.load_workbook(filename, data_only=True )
       # sheet=workbook[sheetName]
        read_row=3
        #read_column=5
        
        self.name=sheet.cell(row=read_row,column=read_column).value#имя секции
        self.name_founding=sheet.cell(row=read_row+1,column=read_column).value
        self.name_wall=sheet.cell(row=read_row+3,column=read_column).value
        self.count=sheet.cell(row=read_row+5,column=read_column).value
        self.foundation_base=sheet.cell(row=read_row+6,column=read_column).value
        self.leght=sheet.cell(row=read_row+7,column=read_column).value*1000
        self.height_start=sheet.cell(row=read_row+8,column=read_column).value*1000
        self.height_end=sheet.cell(row=read_row+9,column=read_column).value*1000
        self.foundation_width=sheet.cell(row=read_row+10,column=read_column).value*1000
        self.top_wall_width=sheet.cell(row=read_row+11,column=read_column).value*1000
        self.edge_distance=sheet.cell(row=read_row+12,column=read_column).value*1000
        self.bottom_wall_width=sheet.cell(row=read_row+13,column=read_column).value*1000
        self.t1=sheet.cell(row=read_row+14,column=read_column).value*1000
        self.t2=sheet.cell(row=read_row+15,column=read_column).value*1000
        self.t3=sheet.cell(row=read_row+16,column=read_column).value*1000
        self.t4=sheet.cell(row=read_row+17,column=read_column).value*1000
        self.V1=sheet.cell(row=read_row+18,column=read_column).value
        self.V2=sheet.cell(row=read_row+19,column=read_column).value

        #
        self.tek_start_x = (sheet.cell(row=read_row+25,column=read_column).value+sheet.cell(row=read_row+27,column=read_column).value)/2
        self.tek_start_y = (sheet.cell(row=read_row+26,column=read_column).value+sheet.cell(row=read_row+28,column=read_column).value)/2
        self.tek_end_x = (sheet.cell(row=read_row+29,column=read_column).value+sheet.cell(row=read_row+31,column=read_column).value)/2
        self.tek_end_y = (sheet.cell(row=read_row+30,column=read_column).value+sheet.cell(row=read_row+32,column=read_column).value)/2



    def ShowAll(self):
        print("Имя секции: "+self.name)
        print("Имя ростверка: "+self.name_founding)
        print("Имя стенки: "+self.name_wall)
        print("Отметка подошвы: "+str(self.foundation_base))
        print("Колличество секций: "+str(self.count))
        print("Длина: "+str(self.leght))
        print("Общая высота (нач): "+str(self.height_start))
        print("Общая высота (кон): "+str(self.height_end))
        print("Ширина фундамента: "+str(self.foundation_width))
        print("Ширина стены сверху : "+str(self.top_wall_width))
        print("Расстояние от стены до края: "+str(self.edge_distance))
        print("Ширина стены внизу: "+str(self.bottom_wall_width))
        print("Толщина перекрытия 1 у стены: "+str(self.t1))
        print("Толщина перекрытия 1 у насыпи: "+str(self.t2))
        print("Толщина перекрытия 2 у стены: "+str(self.t3))
        print("Толщина перекрытия 2 у насыпи: "+str(self.t4))
        print("Объем ростверка подпорной стены, м3: "+str(self.V1))
        print("Объем тела подпорной стены, м3: "+str(self.V2))

        

class Walls(object):
    name: str # Название сооружения
    total_count: int # Колличество секций в сооружении
    sections=[] # Список классов(Wall) параметров секции
    sections_coors=[] # Список списков координат секций 
    
    





