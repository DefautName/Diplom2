from Classes import Wall,Walls
import Functions
import openpyxl
import AutocadWork



# Описание файла и поожения данных для выгрузки в питон
file_Name="DATA_.xlsm" #"C:/Users/Ponka/Desktop/Диплом/Diplom2/DATA_.xlsm"
sheet_Name="_ПС"
read_row=3 #Стартовая строка
read_column=5 #Стартовый столбец считывания данных по секции
workbook = openpyxl.load_workbook(file_Name, data_only=True )
sheet=workbook[sheet_Name]

input_datas=Walls()
input_datas.name=sheet.cell(row=read_row,column=read_column-1)
input_datas.total_count=0
print("Начало считывания данных с Ексель")
while sheet.cell(row=read_row,column=read_column).value != None:#"None"
    input_data = Wall()
    input_data.SetData(sheet,read_column)
    input_datas.total_count += 1
    input_datas.sections.append(input_data)
    read_column+=1
 
print("Отрисовать подпорные стенки?\n1. В Autocad \n2. В Tekla Structures")
choice=int(input())
if choice == 1:
    AutocadWork.DrawAutocad(input_datas)

if choice == 2:
    # try:
    import TeklaWork
    TeklaWork.InputIKomponent(input_datas)
    # except:
    #     print("На компьютере отсутствует Текла или в папке нет одного из его модулей")

input("Нажмите Enter для завершения программы")

